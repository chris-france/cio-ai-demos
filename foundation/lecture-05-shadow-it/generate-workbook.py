"""Generate a realistic complex Excel workbook for the Shadow IT demo.

Run once: python generate-workbook.py
Creates: sample-workbook.xlsx

This workbook simulates a finance team's 'mission-critical spreadsheet' with
deliberate complexity problems a CIO should worry about.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, numbers
from openpyxl.utils import get_column_letter


def create_workbook():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Revenue Forecast ──
    ws1 = wb.active
    ws1.title = "Revenue Forecast"
    ws1.sheet_properties.tabColor = "1F4E79"

    headers = ["Month", "Product A", "Product B", "Product C", "Total", "YoY Growth", "Cumulative"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", fill_type="solid")

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    base_a = [120000, 135000, 142000, 128000, 155000, 168000,
              145000, 152000, 178000, 192000, 185000, 210000]
    base_b = [85000, 92000, 78000, 95000, 88000, 102000,
              97000, 105000, 112000, 98000, 115000, 125000]
    base_c = [45000, 52000, 48000, 55000, 62000, 58000,
              67000, 72000, 68000, 75000, 82000, 95000]

    for i, month in enumerate(months):
        row = i + 2
        ws1.cell(row=row, column=1, value=month)
        ws1.cell(row=row, column=2, value=base_a[i])
        ws1.cell(row=row, column=3, value=base_b[i])
        ws1.cell(row=row, column=4, value=base_c[i])
        # Total — normal SUM
        ws1.cell(row=row, column=5).value = f"=SUM(B{row}:D{row})"
        # YoY Growth — HARDCODED percentages (bad practice)
        growth_vals = [0.12, 0.15, 0.08, 0.18, 0.22, 0.14, 0.11, 0.19, 0.25, 0.16, 0.13, 0.21]
        ws1.cell(row=row, column=6, value=growth_vals[i])
        ws1.cell(row=row, column=6).number_format = '0.0%'
        # Cumulative — references previous row (creates chain dependency)
        if i == 0:
            ws1.cell(row=row, column=7).value = f"=E{row}"
        else:
            ws1.cell(row=row, column=7).value = f"=G{row-1}+E{row}"

    # Totals row
    row = 14
    ws1.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    for col in range(2, 6):
        letter = get_column_letter(col)
        ws1.cell(row=row, column=col).value = f"=SUM({letter}2:{letter}13)"
        ws1.cell(row=row, column=col).font = Font(bold=True)

    # ── Sheet 2: Cost Allocation ──
    ws2 = wb.create_sheet("Cost Allocation")
    ws2.sheet_properties.tabColor = "C00000"

    headers = ["Department", "Headcount", "Avg Salary", "Benefits Rate",
               "Total Comp", "Overhead %", "Fully Loaded", "Revenue Share"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C00000", fill_type="solid")

    depts = [
        ("Engineering", 45, 142000, 0.32),
        ("Sales", 28, 95000, 0.28),
        ("Marketing", 15, 88000, 0.25),
        ("Finance", 8, 105000, 0.30),
        ("HR", 6, 92000, 0.28),
        ("Operations", 18, 85000, 0.25),
        ("Executive", 4, 225000, 0.35),
    ]

    for i, (dept, hc, sal, ben) in enumerate(depts):
        row = i + 2
        ws2.cell(row=row, column=1, value=dept)
        ws2.cell(row=row, column=2, value=hc)
        ws2.cell(row=row, column=3, value=sal)
        ws2.cell(row=row, column=4, value=ben)
        ws2.cell(row=row, column=4).number_format = '0.0%'
        # Total comp = headcount * salary * (1 + benefits)
        ws2.cell(row=row, column=5).value = f"=B{row}*C{row}*(1+D{row})"
        # Overhead % — HARDCODED (should be calculated from actuals)
        overhead = [0.15, 0.22, 0.18, 0.10, 0.08, 0.20, 0.05]
        ws2.cell(row=row, column=6, value=overhead[i])
        ws2.cell(row=row, column=6).number_format = '0.0%'
        # Fully loaded = total comp * (1 + overhead)
        ws2.cell(row=row, column=7).value = f"=E{row}*(1+F{row})"
        # Revenue share — CROSS-SHEET reference to Revenue Forecast total
        ws2.cell(row=row, column=8).value = f"=G{row}/'Revenue Forecast'!E14"
        ws2.cell(row=row, column=8).number_format = '0.0%'

    # ── Sheet 3: Vendor Contracts ──
    ws3 = wb.create_sheet("Vendor Contracts")
    ws3.sheet_properties.tabColor = "2E75B6"

    headers = ["Vendor", "Annual Cost", "Users", "Cost/User", "Contract End",
               "Auto-Renew", "Utilization", "Risk Score"]
    for col, h in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E75B6", fill_type="solid")

    vendors = [
        ("Salesforce", 285000, 28, "2025-12-31", "Yes", 0.72),
        ("Microsoft 365", 156000, 124, "2026-03-15", "Yes", 0.91),
        ("AWS", 480000, 45, "2025-06-30", "No", 0.65),
        ("Slack", 42000, 124, "2025-09-30", "Yes", 0.88),
        ("Jira", 36000, 65, "2026-01-31", "Yes", 0.78),
        ("Zoom", 28000, 124, "2025-08-15", "Yes", 0.45),
        ("HubSpot", 72000, 15, "2025-11-30", "Yes", 0.52),
        ("Datadog", 96000, 12, "2026-02-28", "No", 0.83),
        ("DocuSign", 18000, 35, "2025-07-31", "Yes", 0.35),
        ("Figma", 24000, 8, "2025-10-31", "Yes", 0.90),
    ]

    for i, (vendor, cost, users, end, renew, util) in enumerate(vendors):
        row = i + 2
        ws3.cell(row=row, column=1, value=vendor)
        ws3.cell(row=row, column=2, value=cost)
        ws3.cell(row=row, column=3, value=users)
        # Cost per user — formula
        ws3.cell(row=row, column=4).value = f"=B{row}/C{row}"
        ws3.cell(row=row, column=4).number_format = '$#,##0'
        ws3.cell(row=row, column=5, value=end)
        ws3.cell(row=row, column=6, value=renew)
        ws3.cell(row=row, column=7, value=util)
        ws3.cell(row=row, column=7).number_format = '0%'
        # Risk score — nested IF formula (the kind that breaks)
        ws3.cell(row=row, column=8).value = (
            f'=IF(AND(G{row}<0.5,F{row}="Yes"),"HIGH",'
            f'IF(G{row}<0.5,"MEDIUM",'
            f'IF(AND(G{row}>0.8,B{row}<50000),"LOW","MEDIUM")))'
        )

    # ── Sheet 4: Assumptions (hidden danger) ──
    ws4 = wb.create_sheet("Assumptions")
    ws4.sheet_properties.tabColor = "FFC000"

    ws4.cell(row=1, column=1, value="ASSUMPTIONS — DO NOT MODIFY").font = Font(
        bold=True, color="FF0000", size=14
    )
    ws4.cell(row=3, column=1, value="Parameter")
    ws4.cell(row=3, column=2, value="Value")
    ws4.cell(row=3, column=3, value="Last Updated")
    ws4.cell(row=3, column=4, value="Updated By")

    assumptions = [
        ("Tax Rate", 0.21, "2022-01-15", "jsmith"),
        ("Discount Rate", 0.08, "2021-06-30", "UNKNOWN"),
        ("Growth Target", 0.15, "2023-03-01", "cfo_direct"),
        ("Inflation Adj", 0.035, "2020-12-01", "UNKNOWN"),
        ("FX Rate EUR/USD", 1.08, "2024-02-15", "treasury"),
        ("Benefits Load", 0.30, "2019-08-22", "hr_admin"),
        ("Contingency %", 0.10, "2021-11-01", "UNKNOWN"),
    ]

    for i, (param, val, date, by) in enumerate(assumptions):
        row = i + 4
        ws4.cell(row=row, column=1, value=param)
        ws4.cell(row=row, column=2, value=val)
        ws4.cell(row=row, column=3, value=date)
        ws4.cell(row=row, column=4, value=by)

    # ── Sheet 5: Pivot Summary (circular reference setup) ──
    ws5 = wb.create_sheet("Summary")
    ws5.sheet_properties.tabColor = "548235"

    ws5.cell(row=1, column=1, value="EXECUTIVE SUMMARY").font = Font(bold=True, size=14)
    ws5.cell(row=3, column=1, value="Total Revenue")
    ws5.cell(row=3, column=2).value = "='Revenue Forecast'!E14"
    ws5.cell(row=4, column=1, value="Total Costs")
    ws5.cell(row=4, column=2).value = "=SUM('Cost Allocation'!G2:G8)"
    ws5.cell(row=5, column=1, value="Gross Margin")
    ws5.cell(row=5, column=2).value = "=B3-B4"
    ws5.cell(row=6, column=1, value="Margin %")
    ws5.cell(row=6, column=2).value = "=B5/B3"
    ws5.cell(row=6, column=2).number_format = '0.0%'

    ws5.cell(row=8, column=1, value="Vendor Spend")
    ws5.cell(row=8, column=2).value = "=SUM('Vendor Contracts'!B2:B11)"
    ws5.cell(row=9, column=1, value="Vendor % of Revenue")
    ws5.cell(row=9, column=2).value = "=B8/B3"
    ws5.cell(row=9, column=2).number_format = '0.0%'

    # Circular-ish reference: Summary references Cost Allocation which references Revenue Forecast
    # which... a student exploring with CC will discover the dependency chain
    ws5.cell(row=11, column=1, value="Adjusted Margin (after assumptions)")
    ws5.cell(row=11, column=2).value = "=B5*(1-Assumptions!B7)*(1+Assumptions!B4)"

    # Format currency columns
    for ws in [ws1, ws2, ws3, ws5]:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value > 1000:
                    cell.number_format = '$#,##0'

    wb.save("sample-workbook.xlsx")
    print("Created sample-workbook.xlsx")
    print("Sheets: Revenue Forecast, Cost Allocation, Vendor Contracts, Assumptions, Summary")
    print("Complexity: cross-sheet references, hardcoded values, stale assumptions, nested IFs")


if __name__ == "__main__":
    create_workbook()
